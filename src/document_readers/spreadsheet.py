import openpyxl
import xlrd
from pathlib import Path
from typing import Callable


def read_excel(
    path: str,
    filter_fn: Callable[[dict], bool] = lambda row: True,
    skip_first_row: bool = True
) -> list[dict]:
    """
    Reads an Excel spreadsheet and returns filtered rows as a list of dicts.

    Args:
        path:           Path to the .xls or .xlsx file.
        filter_fn:      A predicate that receives a row dict and returns True
                        to include it, False to exclude it. Defaults to
                        including all rows.
        skip_first_row: If True (default), the first row is excluded from results.

    Returns:
        A list of dicts where each key is a column header (or integer index
        if skip_first_row is False) and each value is the corresponding cell
        value for that row.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError:        If the file extension is not .xls or .xlsx.
    """
    file = Path(path)

    if not file.exists():
        raise FileNotFoundError(f"File not found: {path}")

    match file.suffix.lower():
        case ".xlsx":
            return _read_xlsx(file, filter_fn, skip_first_row)
        case ".xls":
            return _read_xls(file, filter_fn, skip_first_row)
        case _:
            raise ValueError(f"Unsupported file type '{file.suffix}'. Expected .xls or .xlsx.")


def _read_xlsx(file: Path, filter_fn: Callable[[dict], bool], skip_first_row: bool) -> list[dict]:
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    if skip_first_row:
        headers = list(next(rows))
    else:
        headers = list(range(ws.max_column))

    result = [
        row_dict
        for row in rows
        if filter_fn(row_dict := dict(zip(headers, row)))
    ]

    wb.close()
    return result


def _read_xls(file: Path, filter_fn: Callable[[dict], bool], skip_first_row: bool) -> list[dict]:
    wb = xlrd.open_workbook(file)
    ws = wb.sheet_by_index(0)

    if skip_first_row:
        headers = [ws.cell_value(0, col) for col in range(ws.ncols)]
        start_row = 1
    else:
        headers = list(range(ws.ncols))
        start_row = 0

    result = [
        row_dict
        for row_idx in range(start_row, ws.nrows)
        if filter_fn(row_dict := {
            headers[col]: ws.cell_value(row_idx, col)
            for col in range(ws.ncols)
        })
    ]

    return result